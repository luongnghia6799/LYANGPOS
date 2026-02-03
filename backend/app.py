import os
import ctypes

# import openpyxl # Lazy import
# from openpyxl import Workbook


import io
import sys
import webbrowser
import json
from flask import Flask, request, jsonify, send_file, send_from_directory, redirect
from flask_cors import CORS
from models import db, Product, Partner, Order, OrderDetail, CashVoucher, CustomerPrice, AppSetting, ComboItem, PrintTemplate, User, BankAccount, BankTransaction, VoiceAlias
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload
from datetime import datetime
import re
import unicodedata
from sqlalchemy import event, inspect
from sqlalchemy.engine import Engine
from werkzeug.security import generate_password_hash, check_password_hash

def remove_accents(s):
    if not s: return ""
    s = str(s).lower()
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', s)
    s = re.sub(r'[ìíịỉĩ]', 'i', s)
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', s)
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', s)
    s = re.sub(r'[ỳýỵỷỹ]', 'y', s)
    s = re.sub(r'[đ]', 'd', s)
    # Combine with normalization for other characters
    nfkd_form = unicodedata.normalize('NFKD', s)
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)])

def normalize_date_sqlite(date_str):
    if not date_str: return "9999-12-31"
    date_str = str(date_str).strip()
    try:
        if '/' in date_str:
            parts = date_str.split('/')
            if len(parts) == 3:
                day = parts[0].zfill(2)
                month = parts[1].zfill(2)
                year = parts[2]
                if len(year) == 2: year = "20" + year
                return f"{year}-{month}-{day}"
        if '-' in date_str:
            parts = date_str.split('-')
            if len(parts) == 3:
                if len(parts[0]) == 4: return date_str # YYYY-MM-DD
                day = parts[0].zfill(2)
                month = parts[1].zfill(2)
                year = parts[2]
                if len(year) == 2: year = "20" + year
                return f"{year}-{month}-{day}"
    except:
        pass
    return date_str

@event.listens_for(Engine, "connect")
def set_sqlite_custom_func(dbapi_connection, connection_record):
    # Fix: Only register these functions for SQLite connections
    # When using PostgreSQL (Render/Supabase), this would cause an AttributeError
    cursor = dbapi_connection.cursor()
    try:
        # Check if we are really on SQLite
        # SQLite cursors don't usually fail on execute("SELECT sqlite_version()")
        # But a safer way is checking the connection class name or just try/except
        if type(dbapi_connection).__name__ == 'Connection':  # Standard sqlite3 connection name
             dbapi_connection.create_function("remove_accents", 1, remove_accents)
             dbapi_connection.create_function("normalize_date", 1, normalize_date_sqlite)
    except Exception:
        # If it fails (e.g. Postgres connection object doesn't have create_function), just ignore
        pass
    finally:
        cursor.close()

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
        if os.path.exists(os.path.join(base_dir, '_internal')):
             return os.path.join(base_dir, '_internal', relative_path)
        return os.path.join(base_dir, relative_path)
    # Go up one level from backend/ to project root
    base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

# Determine base path safely
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    try:
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    except NameError:
        BASE_DIR = os.getcwd()

def get_storage_path(relative_path):
    return os.path.join(BASE_DIR, relative_path)


static_folder_path = resource_path(os.path.join('frontend', 'dist'))
app = Flask(__name__, static_folder=static_folder_path, static_url_path='/')
app.secret_key = os.environ.get('SECRET_KEY', 'dev_key_super_secret_change_me_in_prod')

# Handle PyInstaller Splash Screen
try:
    import pyi_splash
    # We will close it after server is fully ready in the wait_for_server function
except Exception:
    pass


try:
    myappid = 'com.lyang.pos.v3.farmer.icon.v14' # increment version to refresh taskbar icon with new logo
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
except Exception:
    pass

# Note: User requested 671999 but ports are limited to 65535.
# Setting to 6719 as a high-range alternative.
CURRENT_PORT = 3579

# Logging setup for production
import logging
LOG_FILE = get_storage_path("app_debug.log")
logging.basicConfig(filename=LOG_FILE, level=logging.INFO, 
                    format='%(asctime)s %(levelname)s: %(message)s')
app.logger.info("LyangPOS Start")

# Database Configuration
# Priority: Env Var (Render/Supabase) > Local SQLite
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Fix for SQLAlchemy: Render provides 'postgres://', SQLAlchemy needs 'postgresql://'
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{get_storage_path(os.path.join("instance", "easypos.db"))}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# CORS Configuration: Allow all for dev, restrict in prod via env if needed
# For now, we keep it permissive but ready for restriction
allowed_origins = os.environ.get('ALLOWED_ORIGINS', '*').split(',')
CORS(app, resources={r"/api/*": {"origins": allowed_origins}})

os.makedirs(get_storage_path("instance"), exist_ok=True)
db.init_app(app)
with app.app_context():
    db.create_all()

import time
import threading
LAST_HEARTBEAT = time.time()

@app.route('/api/heartbeat', methods=['POST'])
def heartbeat():
    global LAST_HEARTBEAT
    LAST_HEARTBEAT = time.time()
    return jsonify({'status': 'ok'})

def monitor_heartbeat():
    while True:
        time.sleep(5)
        # If no heartbeat for more than 300 seconds, exit
        if time.time() - LAST_HEARTBEAT > 300:
            app.logger.info("No heartbeat for 300s. Shutting down...")
            os._exit(0)

# Only start monitor if frozen (production)
# Disabled auto-shutdown heartbeat monitor because user wants it to never turn off automatically.
# if getattr(sys, 'frozen', False):
#     heartbeat_thread = threading.Thread(target=monitor_heartbeat)
#     heartbeat_thread.daemon = True
#     heartbeat_thread.start()

import tkinter as tk
from tkinter import ttk

class SplashScreen:
    def __init__(self, port):
        # Safety check for headless environments
        if os.environ.get('NO_GUI') or os.environ.get('HEADLESS'):
            return

        self.port = port
        try:
            self.root = tk.Tk()
        except Exception as e:
            # If Tkinter cannot connect to display (headless), we just abort splash
            print(f"Splash screen skipped: {e}")
            self.root = None
            return

        self.root.title("LyangPOS")
        
        # Try to set icon
        try:
            icon_path = resource_path('frontend/public/favicon.ico')
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except:
            pass

        # Borderless window
        self.root.overrideredirect(True)
        self.root.attributes("-topmost", True)
        
        # Full screen background or large window
        width = 800
        height = 500
        
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
        # Colors
        self.bg_color = '#ecfdf5' # Light Emerald/Mint background
        self.text_color = '#065f46' # Dark Emerald
        self.sub_color = '#34d399' # Light Green for spinner
        self.spinner_color = '#10b981' # Emerald for active spinner part
        
        self.root.configure(bg=self.bg_color)
        
        # Main Canvas
        self.canvas = tk.Canvas(self.root, width=width, height=height, bg=self.bg_color, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
        
        # Center coordinates
        self.cx = width // 2
        self.cy = height // 2 - 30
        
        # Spinner state
        self.angle = 0
        
        # Draw static elements
        # Title
        self.canvas.create_text(self.cx, self.cy + 80, text="LyangPOS", font=("Segoe UI", 24, "bold"), fill=self.text_color)
        
        # Status text
        self.status_text_id = self.canvas.create_text(self.cx, self.cy + 120, text="Đang khởi động hệ thống...", font=("Segoe UI", 10), fill="#6ee7b7")
        
        # Start animation
        self.animate()
        
        # Check server
        self.root.after(1000, self.check_server)
        
    def animate(self):
        self.canvas.delete("spinner")
        
        # Draw background ring
        radius = 40
        self.canvas.create_oval(self.cx - radius, self.cy - radius, self.cx + radius, self.cy + radius, outline="#a7f3d0", width=4, tags="spinner")
        
        # Draw rotating arc
        start = self.angle
        extent = 120
        self.canvas.create_arc(self.cx - radius, self.cy - radius, self.cx + radius, self.cy + radius, start=start, extent=extent, outline=self.spinner_color, width=4, style="arc", tags="spinner")
        
        self.angle = (self.angle - 10) % 360
        self.root.after(20, self.animate)

    def check_server(self):
        import socket
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(0.3)
                if s.connect_ex(('127.0.0.1', self.port)) == 0:
                    self.set_ready()
                    return
        except:
            pass
        self.root.after(500, self.check_server)

    def set_ready(self):
        self.canvas.itemconfig(self.status_text_id, text="Đã kết nối! Đang vào ứng dụng...")
        
        # Open browser
        self.open_browser()
        
        # AUTO CLOSE after 2 seconds
        self.root.after(2000, self.root.destroy)

    def open_browser(self):
        import webbrowser
        target_url = f"http://localhost:{self.port}"
        try:
            chrome_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe")
            ]
            for path in chrome_paths:
                if os.path.exists(path):
                    import subprocess
                    subprocess.Popen([path, f"--app={target_url}", "--start-maximized"])
                    return
            webbrowser.open(target_url)
        except:
            webbrowser.open(target_url)

    def run(self):
        if self.root:
            self.root.mainloop()

def run_migrations():
    with app.app_context():
        try:
            db.create_all()
            # Ensure BankAccount and BankTransaction tables are created if not exist
            inspector = inspect(db.engine)
            if 'bank_account' not in inspector.get_table_names():
                BankAccount.__table__.create(db.engine)
                app.logger.info("Created bank_account table")
            if 'bank_transaction' not in inspector.get_table_names():
                BankTransaction.__table__.create(db.engine)
                app.logger.info("Created bank_transaction table")

            # Auto-migration for missing columns
            # Auto-migration for missing columns
            # inspector already defined above

            
            # Check product table
            columns = [c['name'] for c in inspector.get_columns('product')]
            
            with db.engine.begin() as conn:
                if 'code' not in columns:
                    conn.execute(db.text('ALTER TABLE product ADD COLUMN code TEXT'))
                    app.logger.info("Added column 'code' to product table")
                
                if 'is_combo' not in columns:
                    conn.execute(db.text('ALTER TABLE product ADD COLUMN is_combo BOOLEAN DEFAULT 0'))
                    app.logger.info("Added column 'is_combo' to product table")
                
                if 'active_ingredient' not in columns:
                    conn.execute(db.text('ALTER TABLE product ADD COLUMN active_ingredient TEXT'))
                    app.logger.info("Added column 'active_ingredient' to product table")
                
                if 'brand' not in columns:
                    conn.execute(db.text('ALTER TABLE product ADD COLUMN brand TEXT'))
                    app.logger.info("Added column 'brand' to product table")
                
                # Check order table
                order_columns = [c['name'] for c in inspector.get_columns('order')]
                if 'display_id' not in order_columns:
                    conn.execute(db.text('ALTER TABLE "order" ADD COLUMN display_id TEXT'))
                    app.logger.info("Added column 'display_id' to order table")
                
                # Check cash_voucher table
                cv_columns = [c['name'] for c in inspector.get_columns('cash_voucher')]
                if 'source' not in cv_columns:
                    conn.execute(db.text('ALTER TABLE cash_voucher ADD COLUMN source TEXT DEFAULT "manual"'))
                if 'order_id' not in cv_columns:
                    conn.execute(db.text('ALTER TABLE cash_voucher ADD COLUMN order_id INTEGER'))
                
                # Cleanup previous deletions if any
                pass
                    
        except Exception as e:
            app.logger.error(f"Error creating/migrating database: {e}")

# Run initial migration
run_migrations()

# --- Auth Routes ---
@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    display_name = data.get('display_name', username)
    role = data.get('role', 'User')

    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400

    if User.query.filter_by(username=username).first():
        return jsonify({'error': 'Username already exists'}), 400

    new_user = User(
        username=username,
        password_hash=generate_password_hash(password),
        display_name=display_name,
        role=role
    )
    db.session.add(new_user)
    db.session.commit()

    return jsonify({'message': 'User registered successfully', 'user': new_user.to_dict()}), 201

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')

    user = User.query.filter_by(username=username).first()

    if user and check_password_hash(user.password_hash, password):
        return jsonify({
            'message': 'Login successful',
            'user': user.to_dict()
        }), 200
    
    return jsonify({'error': 'Invalid username or password'}), 401

@app.route('/api/users', methods=['GET'])
def get_users():
    users = User.query.all()
    return jsonify([u.to_dict() for u in users])


# --- Font Management ---
# Ensure storage folders are outside of PyInstaller's temp directory

FONTS_FOLDER = get_storage_path(os.path.join("uploads", "fonts"))
os.makedirs(FONTS_FOLDER, exist_ok=True)

LOGO_FOLDER = get_storage_path(os.path.join("uploads", "logos"))
os.makedirs(LOGO_FOLDER, exist_ok=True)

@app.route('/uploads/fonts/<path:filename>')
def serve_font(filename):
    return send_from_directory(FONTS_FOLDER, filename)

@app.route('/uploads/logos/<path:filename>')
def serve_logo(filename):
    return send_from_directory(LOGO_FOLDER, filename)

@app.route('/api/fonts', methods=['GET'])
def list_fonts():
    if not os.path.exists(FONTS_FOLDER): return jsonify([])
    fonts = [f for f in os.listdir(FONTS_FOLDER) if f.lower().endswith(('.ttf', '.otf', '.woff', '.woff2'))]
    return jsonify(fonts)

@app.route('/api/fonts', methods=['POST'])
def upload_font():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file:
        filename = file.filename
        file.save(os.path.join(FONTS_FOLDER, filename))
        return jsonify({'message': 'Font uploaded successfully', 'filename': filename})

@app.route('/api/upload-logo', methods=['POST'])
def upload_logo():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file:
        filename = file.filename
        file.save(os.path.join(LOGO_FOLDER, filename))
        return jsonify({'url': f'/uploads/logos/{filename}'})

# --- Products ---
@app.route('/api/products', methods=['GET'])
def get_products():
    search = request.args.get('search', '').lower()
    filter_type = request.args.get('filterType', 'all')
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)
    
    # Critical Fix: Optimize N+1 query for combo_items
    query = Product.query.options(joinedload(Product.combo_items))
    
    if search:
        s_norm = remove_accents(search)
        query = query.filter(
            db.func.remove_accents(Product.name).ilike(f'%{s_norm}%') | 
            Product.code.ilike(f'%{search}%') |
            db.func.remove_accents(db.func.coalesce(Product.active_ingredient, '')).ilike(f'%{s_norm}%') |
            db.func.remove_accents(db.func.coalesce(Product.brand, '')).ilike(f'%{s_norm}%')
        )
    
    brand = request.args.get('brand')
    if brand:
        query = query.filter(Product.brand == brand)
    
    if filter_type == 'out_of_stock':
        query = query.filter(Product.stock <= 0)
    elif filter_type == 'warning':
        query = query.filter(Product.stock < 2 * Product.multiplier)
    elif filter_type == 'expired':
        today = datetime.now().strftime('%Y-%m-%d')
        query = query.filter(Product.expiry_date != None, Product.expiry_date != '', db.func.normalize_date(Product.expiry_date) <= today)
    elif filter_type == 'near_expiry':
        from datetime import timedelta
        today = datetime.now()
        target_date = (today + timedelta(days=60)).strftime('%Y-%m-%d')
        today_str = today.strftime('%Y-%m-%d')
        query = query.filter(
            Product.expiry_date != None, 
            Product.expiry_date != '', 
            db.func.normalize_date(Product.expiry_date) > today_str,
            db.func.normalize_date(Product.expiry_date) <= target_date
        )
    elif filter_type == 'loss':
        query = query.filter(Product.sale_price < Product.cost_price, Product.is_combo == False)

    sort_by = request.args.get('sort_by', 'name')
    sort_order = request.args.get('sort_order', 'asc')

    # Mapping frontend sort keys to DB columns
    sort_map = {
        'id': Product.id,
        'name': Product.name,
        'code': Product.code,
        'unit': Product.unit,
        'cost_price': Product.cost_price,
        'sale_price': Product.sale_price,
        'stock': Product.stock,
        'expiry_date': Product.expiry_date
    }
    
    sort_col = sort_map.get(sort_by, Product.name)
    if sort_order == 'desc':
        query = query.order_by(sort_col.desc())
    else:
        query = query.order_by(sort_col.asc())

    if page and limit:
        pagination = query.paginate(page=page, per_page=limit, error_out=False)
        products = pagination.items
        total = pagination.total
        pages = pagination.pages
        current_page = pagination.page
    else:
        products = query.all()
        total = len(products)
        pages = 1
        current_page = 1

    results = []
    for p in products:
        d = p.to_dict()
        if p.is_combo:
            d['combo_items'] = [i.to_dict() for i in p.combo_items]
        results.append(d)
        
    if page and limit:
        return jsonify({
            'items': results,
            'total': total,
            'pages': pages,
            'current_page': current_page
        })
    else:
        return jsonify(results)

@app.route('/api/products', methods=['POST'])
def create_product():
    data = request.json
    new_prod = Product(
        name=data['name'],
        code=data.get('code'),
        unit=data.get('unit'),
        secondary_unit=data.get('secondary_unit'),
        multiplier=data.get('multiplier', 1),
        cost_price=data.get('cost_price', 0),
        sale_price=data.get('sale_price', 0),
        stock=data.get('stock', 0),
        expiry_date=data.get('expiry_date'),
        active_ingredient=data.get('active_ingredient'),
        brand=data.get('brand'),
        is_combo=data.get('is_combo', False)
    )
    db.session.add(new_prod)
    db.session.flush() # Get ID

    if data.get('is_combo') and 'combo_items' in data:
        for item in data['combo_items']:
            ci = ComboItem(
                combo_id=new_prod.id,
                product_id=item['product_id'],
                quantity=item['quantity']
            )
            db.session.add(ci)

    db.session.commit()
    return jsonify(new_prod.to_dict()), 201

@app.route('/api/products/<int:id>', methods=['PUT'])
def update_product(id):
    prod = Product.query.get_or_404(id)
    data = request.json
    prod.name = data.get('name', prod.name)
    prod.code = data.get('code', prod.code)
    prod.unit = data.get('unit', prod.unit)
    prod.secondary_unit = data.get('secondary_unit', prod.secondary_unit)
    prod.multiplier = data.get('multiplier', prod.multiplier)
    prod.cost_price = data.get('cost_price', prod.cost_price)
    prod.sale_price = data.get('sale_price', prod.sale_price)
    prod.stock = data.get('stock', prod.stock)
    prod.expiry_date = data.get('expiry_date', prod.expiry_date)
    prod.active_ingredient = data.get('active_ingredient', prod.active_ingredient)
    prod.brand = data.get('brand', prod.brand)
    prod.is_combo = data.get('is_combo', prod.is_combo)
    
    if 'combo_items' in data:
        # Update combo items
        ComboItem.query.filter_by(combo_id=prod.id).delete()
        for item in data['combo_items']:
            ci = ComboItem(
                combo_id=prod.id,
                product_id=item['product_id'],
                quantity=item['quantity']
            )
            db.session.add(ci)

    db.session.commit()
    # Include items in response if combo
    res = prod.to_dict()
    if prod.is_combo:
        res['combo_items'] = [i.to_dict() for i in prod.combo_items]
    return jsonify(res)

@app.route('/api/products/<int:id>', methods=['DELETE'])
def delete_product(id):
    # Check if used in transactions
    in_use = OrderDetail.query.filter_by(product_id=id).first()
    if in_use:
        return jsonify({'error': 'Không thể xóa sản phẩm đã có lịch sử giao dịch'}), 400
    
    prod = Product.query.get_or_404(id)
    db.session.delete(prod)
    db.session.commit()
    return jsonify({'message': 'Deleted successfully'})

@app.route('/api/products/brands', methods=['GET'])
def get_product_brands():
    brands = db.session.query(Product.brand).distinct().all()
    # Filter out None or empty strings and sort
    brand_list = sorted([b[0] for b in brands if b[0]])
    return jsonify(brand_list)

@app.route('/api/voice-aliases', methods=['GET'])
def get_voice_aliases():
    try:
        aliases = VoiceAlias.query.all()
        return jsonify([a.to_dict() for a in aliases])
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/voice-aliases', methods=['POST'])
def add_voice_alias():
    data = request.json
    if not data or 'product_id' not in data or 'alias_name' not in data:
        return jsonify({'error': 'Missing data'}), 400
    
    alias = VoiceAlias(
        product_id=data['product_id'],
        alias_name=data['alias_name'].lower().strip()
    )
    db.session.add(alias)
    db.session.commit()
    return jsonify(alias.to_dict()), 201

@app.route('/api/products/bulk-delete', methods=['POST'])
def bulk_delete_products():
    data = request.json
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': 'No IDs provided'}), 400
    
    try:
        # Check for products in use
        in_use_count = OrderDetail.query.filter(OrderDetail.product_id.in_(ids)).group_by(OrderDetail.product_id).count()
        if in_use_count > 0:
            return jsonify({'error': f'Có {in_use_count} sản phẩm đang có lịch sử giao dịch và không thể xóa.'}), 400
        
        # Delete if not in use
        deleted = Product.query.filter(Product.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'message': f'Đã xóa {deleted} sản phẩm thành công'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/products/import', methods=['POST'])
def import_products():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # Get headers
        headers = [str(cell.value).strip() for cell in ws[1]]
        
        def get_val(row, col_name, default=None):
            try:
                idx = headers.index(col_name)
                val = row[idx].value
                return val if val is not None else default
            except (ValueError, IndexError):
                return default

        count = 0
        # Iterate from row 2
        for row in ws.iter_rows(min_row=2):
            name = str(get_val(row, 'Tên sản phẩm', '')).strip()
            if not name or name == 'None' or name == '': continue
            
            prod = Product.query.filter_by(name=name).first()
            if not prod:
                prod = Product(name=name, unit=str(get_val(row, 'Đơn vị', 'Cái')))
                db.session.add(prod)
            
            prod.unit = str(get_val(row, 'Đơn vị', prod.unit))
            sec_unit = get_val(row, 'Đơn vị phụ')
            if sec_unit is not None: prod.secondary_unit = str(sec_unit)
            
            multiplier = get_val(row, 'Quy cách')
            if multiplier is not None: prod.multiplier = float(multiplier)
            
            cost_price = get_val(row, 'Giá vốn')
            if cost_price is not None: prod.cost_price = float(cost_price)
            
            sale_price = get_val(row, 'Giá bán')
            if sale_price is not None: prod.sale_price = float(sale_price)
            
            stock = get_val(row, 'Tồn kho')
            if stock is not None: prod.stock = int(float(stock))
            
            expiry = get_val(row, 'Hạn sử dụng')
            if expiry is not None: prod.expiry_date = str(expiry)
            
            ai = get_val(row, 'Hoạt chất')
            if ai is not None: prod.active_ingredient = str(ai)
            
            brand = get_val(row, 'Hãng')
            if brand is not None: prod.brand = str(brand)
            
            count += 1
            
        db.session.commit()
        return jsonify({'message': f'Imported {count} products successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/products/template', methods=['GET'])
def get_template():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    headers = ['Tên sản phẩm', 'Đơn vị', 'Đơn vị phụ', 'Quy cách', 'Giá vốn', 'Giá bán', 'Tồn kho', 'Hạn sử dụng', 'Hoạt chất', 'Hãng']
    ws.append(headers)
    ws.append(['Sản phẩm mẫu', 'Chai', 'Thùng', 24, 100000, 150000, 10, '31/12/2026', 'Abamectin', 'Lyang Nghĩa'])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='mau_nhap_san_pham.xlsx'
    )


# --- Product Orders ---
@app.route('/api/products/<int:id>/orders', methods=['GET'])
def get_product_orders(id):
    # Find all orders that contain this product
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    
    query = Order.query.join(OrderDetail).filter(OrderDetail.product_id == id)
    
    if year:
        query = query.filter(db.func.strftime('%Y', Order.date) == str(year))
    if month:
        query = query.filter(db.func.strftime('%m', Order.date) == str(month).zfill(2))
    if day:
        query = query.filter(db.func.strftime('%d', Order.date) == str(day).zfill(2))
        
    orders = query.order_by(Order.date.desc()).all()
    return jsonify([o.to_dict() for o in orders])

@app.route('/api/products/<int:id>/history', methods=['GET'])
def get_product_history(id):
    try:
        # Get all movements for this product
        # 1. From Order Details
        details = OrderDetail.query.filter_by(product_id=id).join(Order).order_by(Order.date.desc()).all()
        
        history = []
        for d in details:
            change_qty = 0
            type_label = ''
            
            if d.order.type == 'Sale':
                # Sale means Out
                change_qty = -d.quantity
                type_label = 'Bán hàng'
            elif d.order.type == 'Purchase':
                # Purchase means In
                change_qty = d.quantity
                type_label = 'Nhập hàng'
                
            # Handle Combos if necessary (Combos don't change stock directly? Wait, OrderDetail links to Product. 
            # If product is Combo, its stock decreases. But usually we track components.
            # If this request is for a Component Product, we must find Orders where this Component was part of a Combo.
            
            history.append({
                'date': d.order.date.isoformat(),
                'display_id': d.order.display_id,
                'order_id': d.order.id,
                'partner_name': d.order.partner.name if d.order.partner else ('Khách Lẻ' if d.order.type=='Sale' else 'NCC Vãng Lai'),
                'type': type_label,
                'quantity_change': change_qty,
                'price': d.price
            })

        # Also need to check if this product is part of any Combos that were sold
        # Find all combos that include this product
        parent_combos = ComboItem.query.filter_by(product_id=id).all()
        for ci in parent_combos:
            combo_prod_id = ci.combo_id
            # Find sales of this combo
            combo_details = OrderDetail.query.filter_by(product_id=combo_prod_id).join(Order).all()
            for cd in combo_details:
                qty_deducted = cd.quantity * ci.quantity # Total items used
                
                type_label = 'Bán Combo'
                if cd.order.type == 'Purchase': type_label = 'Nhập Combo' # Rare
                
                # If Sale, we lost stock
                change = -qty_deducted if cd.order.type == 'Sale' else qty_deducted
                
                history.append({
                    'date': cd.order.date.isoformat(),
                    'display_id': cd.order.display_id,
                    'order_id': cd.order.id,
                    'partner_name': cd.order.partner.name if cd.order.partner else 'Khách Lẻ',
                    'type': f"{type_label} ({cd.product.name})",
                    'quantity_change': change,
                    'price': 0 # Hard to attribute price
                })
        
        # Sort combined history by date desc
        history.sort(key=lambda x: x['date'], reverse=True)
        
        return jsonify(history)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# --- Combo Items ---
@app.route('/api/combos/<int:combo_id>/items', methods=['GET'])
def get_combo_items(combo_id):
    items = ComboItem.query.filter_by(combo_id=combo_id).all()
    return jsonify([i.to_dict() for i in items])

@app.route('/api/combos/<int:combo_id>/items', methods=['POST'])
def set_combo_items(combo_id):
    data = request.json # Expected list of {product_id, quantity}
    try:
        # Clear existing
        ComboItem.query.filter_by(combo_id=combo_id).delete()
        for item in data:
            new_item = ComboItem(
                combo_id=combo_id,
                product_id=item['product_id'],
                quantity=item['quantity']
            )
            db.session.add(new_item)
        db.session.commit()
        return jsonify({'message': 'Combo items updated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# --- Partners ---
@app.route('/api/partners', methods=['GET'])
def get_partners():
    search = request.args.get('search', '').lower()
    partner_type = request.args.get('type', 'All')
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)
    
    query = Partner.query
    
    if search:
        s_norm = remove_accents(search)
        query = query.filter(db.func.remove_accents(Partner.name).ilike(f'%{s_norm}%') | Partner.phone.ilike(f'%{search}%'))
    
    if partner_type == 'Customer':
        query = query.filter(Partner.is_customer == True)
    elif partner_type == 'Supplier':
        query = query.filter(Partner.is_supplier == True)
    elif partner_type == 'Both':
        query = query.filter(Partner.is_customer == True, Partner.is_supplier == True)

    sort_by = request.args.get('sort_by', 'name')
    sort_order = request.args.get('sort_order', 'asc')

    sort_map = {
        'name': Partner.name,
        'type': Partner.type,
        'debt_balance': Partner.debt_balance,
        'phone': Partner.phone
    }

    sort_col = sort_map.get(sort_by, Partner.name)
    if sort_order == 'desc':
        query = query.order_by(sort_col.desc())
    else:
        query = query.order_by(sort_col.asc())

    if page and limit:
        pagination = query.paginate(page=page, per_page=limit, error_out=False)
        return jsonify({
            'items': [p.to_dict() for p in pagination.items],
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': pagination.page
        })
    else:
        partners = query.all()
        return jsonify([p.to_dict() for p in partners])

@app.route('/api/partners', methods=['POST'])
def create_partner():
    data = request.json
    new_partner = Partner(
        name=data['name'],
        is_customer=data.get('is_customer', True),
        is_supplier=data.get('is_supplier', False),
        cccd=data.get('cccd'),
        phone=data.get('phone'),
        address=data.get('address'),
        debt_balance=data.get('debt_balance', 0)
    )
    # Sync 'type' for backward compatibility
    if new_partner.is_customer and new_partner.is_supplier:
        new_partner.type = 'Both'
    elif new_partner.is_supplier:
        new_partner.type = 'Supplier'
    else:
        new_partner.type = 'Customer'
        
    db.session.add(new_partner)
    db.session.commit()
    
    # Create Opening Balance Order if debt exists
    if new_partner.debt_balance != 0:
        create_opening_balance_order(new_partner.id, new_partner.debt_balance, new_partner.type)
        
    return jsonify(new_partner.to_dict()), 201

@app.route('/api/partners/<int:id>', methods=['PUT'])
def update_partner(id):
    try:
        partner = Partner.query.get_or_404(id)
        data = request.json
        
        partner.name = data.get('name', partner.name)
        partner.is_customer = data.get('is_customer', partner.is_customer)
        partner.is_supplier = data.get('is_supplier', partner.is_supplier)
        partner.cccd = data.get('cccd', partner.cccd)
        partner.phone = data.get('phone', partner.phone)
        partner.address = data.get('address', partner.address)
        
        # Sync 'type' for backward compatibility
        if partner.is_customer and partner.is_supplier:
            partner.type = 'Both'
        elif partner.is_supplier:
            partner.type = 'Supplier'
        else:
            partner.type = 'Customer'
        
        # Only update debt if it's explicitly provided in the request
        if 'debt_balance' in data:
            partner.debt_balance = float(data['debt_balance'])
            
        db.session.commit()
        return jsonify(partner.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/partners/<int:id>', methods=['DELETE'])
def delete_partner(id):
    try:
        # Check if has orders
        order_count = Order.query.filter_by(partner_id=id).count()
        if order_count > 0:
            return jsonify({'error': f'Không thể xóa đối tác vì đã có {order_count} hóa đơn/giao dịch. Hãy xóa các hóa đơn này trong Lịch sử trước.'}), 400
        
        # Check if has vouchers
        voucher_count = CashVoucher.query.filter_by(partner_id=id).count()
        if voucher_count > 0:
            return jsonify({'error': f'Không thể xóa đối tác vì đã có {voucher_count} phiếu thu/chi. Hãy xóa các phiếu này trong Quỹ tiền trước.'}), 400
        
        partner = Partner.query.get_or_404(id)
        db.session.delete(partner)
        db.session.commit()
        return jsonify({'message': 'Deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/partners/bulk-delete', methods=['POST'])
def bulk_delete_partners():
    data = request.json
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': 'No IDs provided'}), 400
    
    try:
        # Check if any have orders or vouchers
        in_use_orders = Order.query.filter(Order.partner_id.in_(ids)).group_by(Order.partner_id).count()
        in_use_vouchers = CashVoucher.query.filter(CashVoucher.partner_id.in_(ids)).group_by(CashVoucher.partner_id).count()
        
        if in_use_orders > 0 or in_use_vouchers > 0:
            return jsonify({'error': 'Có một số đối tác đã có lịch sử đơn hàng hoặc phiếu thu chi, không thể xóa hàng loạt.'}), 400
            
        deleted = Partner.query.filter(Partner.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'message': f'Đã xóa {deleted} đối tác thành công'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/partners/import', methods=['POST'])
def import_partners():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        headers = [str(cell.value).strip() for cell in ws[1]]
        
        def get_val(row, col_name, default=None):
            try:
                idx = headers.index(col_name)
                val = row[idx].value
                return val if val is not None else default
            except (ValueError, IndexError):
                return default

        count = 0
        for row in ws.iter_rows(min_row=2):
            name = str(get_val(row, 'Tên đối tác', '')).strip()
            if not name or name == 'None' or name == '': continue
            
            p_type_raw = str(get_val(row, 'Loại', 'Customer')).strip()
            # Normalize type
            if 'khách' in p_type_raw.lower() and 'ncc' in p_type_raw.lower(): p_type = 'Both'
            elif 'khách' in p_type_raw.lower(): p_type = 'Customer'
            elif 'ncc' in p_type_raw.lower() or 'nhà cung cấp' in p_type_raw.lower(): p_type = 'Supplier'
            else: p_type = 'Customer'
            
            # Determine flags
            is_cust = True
            is_supp = False
            if p_type == 'Both':
                is_cust = True
                is_supp = True
            elif p_type == 'Supplier':
                is_cust = False
                is_supp = True
            elif p_type == 'Customer':
                is_cust = True
                is_supp = False

            partner = Partner.query.filter_by(name=name, type=p_type).first()
            if not partner:
                partner = Partner(name=name, type=p_type, is_customer=is_cust, is_supplier=is_supp)
                db.session.add(partner)
            else:
                # Update existing if needed
                partner.is_customer = is_cust
                partner.is_supplier = is_supp
            
            phone = get_val(row, 'Số điện thoại')
            if phone is not None: partner.phone = str(phone)
            
            address = get_val(row, 'Địa chỉ')
            if address is not None: partner.address = str(address)
            
            debt = get_val(row, 'Nợ đầu kỳ')
            if debt is not None: 
                d_val = float(debt)
                partner.debt_balance = d_val
                # Add to session first, flush to get ID, then we will handle opening order
                # But to handle it cleanly, wait until after main loop?
                # Actually, we can just create the Order object here and link it.
                # Since we haven't committed partner yet, we can't link validation might fail? 
                # No, SQLAlchemy handles uncommitted objects in session.
                # However, to avoid complexity, let's just flush.
                db.session.flush() # Ensure partner.id is set
                
                # Check if 'NODAU' exists
                existing_nodau = Order.query.filter_by(partner_id=partner.id, display_id='NODAU').first()
                if existing_nodau:
                    # Update it
                    is_positive = d_val > 0
                    existing_nodau.type = 'Sale' if is_positive else 'Purchase'
                    existing_nodau.total_amount = abs(d_val)
                elif d_val != 0:
                    # Create new
                    is_positive = d_val > 0
                    order_type = 'Sale' if is_positive else 'Purchase'
                    new_nodau = Order(
                        partner_id=partner.id,
                        type=order_type,
                        payment_method='Debt',
                        display_id='NODAU',
                        total_amount=abs(d_val),
                        note='Nợ đầu kỳ (Import)',
                        amount_paid=0,
                        date=datetime.now()
                    )
                    db.session.add(new_nodau)
            
            count += 1
            
        db.session.commit()
        
        return jsonify({'message': f'Đã nhập {count} đối tác thành công! Lịch sử công nợ đầu kỳ đã được ghi nhận.'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

def create_opening_balance_order(partner_id, amount, p_type=None):
    # Determine type based on amount and partner type
    # If Amount > 0: Customer owes us (Sale/Debt)
    # If Amount < 0: We owe Supplier (Purchase/Debt)
    
    # However, standard convention:
    # Sale (Debt) adds to debt_balance (+)
    # Purchase (Debt) subtracts from debt_balance (-)
    
    is_positive = amount > 0
    order_type = 'Sale' if is_positive else 'Purchase'
    abs_amount = abs(amount)
    
    order = Order(
        partner_id=partner_id,
        type=order_type,
        payment_method='Debt',
        display_id='#NODAU',
        total_amount=abs_amount,
        note='Nợ đầu kỳ',
        amount_paid=0,
        date=datetime.now() # Should strictly consist with creation, but now is fine
    )
    db.session.add(order)
    db.session.commit()

@app.route('/api/partners/<int:id>/fix-opening-balance', methods=['POST'])
def fix_opening_balance(id):
    partner = Partner.query.get_or_404(id)
    data = request.json
    amount = data.get('amount') # The missing amount to record
    
    if not amount:
        return jsonify({'error': 'Amount required'}), 400
        
    try:
        create_opening_balance_order(id, float(amount))
        return jsonify({'message': 'Đã ghi nhận nợ đầu kỳ thành công'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/partners/template', methods=['GET'])
def get_partner_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Partners"
    
    headers = ['Tên đối tác', 'Loại', 'Số điện thoại', 'Địa chỉ', 'Nợ đầu kỳ']
    ws.append(headers)
    ws.append(['Nguyễn Văn A', 'Khách hàng', '0901234567', 'Hà Nội', 0])
    ws.append(['Công ty X', 'NCC', '0281234567', 'TP HCM', 1500000])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='mau_nhap_doi_tac.xlsx'
    )

@app.route('/api/partners/<int:id>/debt-cycles', methods=['GET'])
def get_partner_debt_cycles(id):
    # Fetch all orders and vouchers for the partner
    orders = Order.query.filter_by(partner_id=id).order_by(Order.date.asc()).all()
    vouchers = CashVoucher.query.filter_by(partner_id=id).order_by(CashVoucher.date.asc()).all()
    
    # Combine and sort by date
    timeline = []
    for o in orders:
        timeline.append({'type': 'order', 'date': o.date, 'obj': o})
    for v in vouchers:
        timeline.append({'type': 'voucher', 'date': v.date, 'obj': v})
    
    timeline.sort(key=lambda x: x['date'])
    
    cycles = []
    current_cycle = None
    balance = 0
    
    for item in timeline:
        if item['type'] == 'order':
            o = item['obj']
            # Debt increases based on type
            # Sale: total - paid = debt to me (+)
            # Purchase: total - paid = my debt to them (-)
            # But let's simplify: absolute debt is what matters for "period"
            # Actually, user says debt period starts with first debt order.
            
            # IMPORTANT: For Debt orders, we add the FULL total_amount to the balance.
            # We ignore amount_paid here because the payment is recorded 
            # via a separate Voucher by the Settlement/POS logic.
            # This ensures symmetry (+Goods, -Payment) and avoids double-counting.
            if o.payment_method == 'Debt':
                if o.type == 'Sale':
                    balance += o.total_amount
                else:
                    balance -= o.total_amount
                
            if o.payment_method == 'Debt' and current_cycle is None and balance != 0:
                current_cycle = {
                    'start_date': o.date.isoformat(),
                    'end_date': None,
                    'status': 'Đang còn nợ'
                }
        else:
            v = item['obj']
            if v.type == 'Receipt': # Customer pays us
                balance -= v.amount
            else: # We pay supplier
                balance += v.amount
        
        # Use a small epsilon for float comparison
        if current_cycle and abs(balance) < 1:
            current_cycle['end_date'] = item['date'].isoformat()
            current_cycle['status'] = 'Đã tất toán'
            cycles.append(current_cycle)
            current_cycle = None
            
    if current_cycle:
        cycles.append(current_cycle)
        
    return jsonify(cycles)

@app.route('/api/partners/<int:id>/recalculate-debt', methods=['POST'])
def recalculate_partner_debt(id):
    partner = Partner.query.get_or_404(id)
    try:
        # Sum of all Debt orders
        # Sale (+) / Purchase (-)
        sale_debt = db.session.query(db.func.sum(Order.total_amount)).filter(Order.partner_id == id, Order.payment_method == 'Debt', Order.type == 'Sale').scalar() or 0
        purchase_debt = db.session.query(db.func.sum(Order.total_amount)).filter(Order.partner_id == id, Order.payment_method == 'Debt', Order.type == 'Purchase').scalar() or 0
        
        # Sum of all vouchers
        # Receipt (-) : Customer pays us
        # Payment (+) : We pay supplier
        receipts = db.session.query(db.func.sum(CashVoucher.amount)).filter(CashVoucher.partner_id == id, CashVoucher.type == 'Receipt').scalar() or 0
        payments = db.session.query(db.func.sum(CashVoucher.amount)).filter(CashVoucher.partner_id == id, CashVoucher.type == 'Payment').scalar() or 0
        
        new_balance = (sale_debt - purchase_debt) - (receipts - payments)
        partner.debt_balance = new_balance
        db.session.commit()
        return jsonify({'message': 'Recalculated successfully', 'new_balance': new_balance})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/partners/<int:id>/ledger', methods=['GET'])
def get_partner_ledger(id):
    # Retrieve full ledger for partner
    try:
        partner = Partner.query.get_or_404(id)
        
        # 1. Get all Debt Orders
        orders = Order.query.filter(
            Order.partner_id == id, 
            Order.payment_method == 'Debt'
        ).all()
        
        # 2. Get all Vouchers
        vouchers = CashVoucher.query.filter_by(partner_id=id).all()
        
        # 3. Combine into timeline
        timeline = []
        for o in orders:
            timeline.append({
                'date': o.date,
                'type': 'Order',
                'ref_id': o.display_id,
                'desc': f"{'Mua hàng' if o.type == 'Sale' else 'Nhập hàng'} (Đơn {o.display_id})",
                'obj': o,
                'timestamp': o.date.timestamp()
            })
            
        for v in vouchers:
            desc = v.note
            if not desc:
                if v.type == 'Receipt': desc = 'Phiếu thu tiền'
                else: desc = 'Phiếu chi tiền'
            
            timeline.append({
                'date': v.date,
                'type': 'Voucher',
                'ref_id': f"PT-{v.id}" if v.type == 'Receipt' else f"PC-{v.id}",
                'desc': desc,
                'obj': v,
                'timestamp': v.date.timestamp()
            })
            
        # Sort by Date Ascending for Calculation
        timeline.sort(key=lambda x: x['date'])
        
        ledger = []
        balance = 0
        
        for item in timeline:
            row = {
                'date': item['date'].isoformat(),
                'ref_id': item['ref_id'],
                'desc': item['desc'],
                'type': item['type'],
                'increase': 0,
                'decrease': 0,
                'running_balance': 0,
                'obj': item['obj'].to_dict()
            }
            
            if item['type'] == 'Order':
                o = item['obj']
                if o.type == 'Sale':
                    row['increase'] = o.total_amount
                    balance += o.total_amount
                else:
                    row['decrease'] = o.total_amount
                    balance -= o.total_amount
            else:
                v = item['obj']
                if v.type == 'Receipt':
                    row['decrease'] = v.amount
                    balance -= v.amount
                else: 
                    row['increase'] = v.amount
                    balance += v.amount
            
            row['running_balance'] = balance
            ledger.append(row)
            
        # For display, reverse to show Newest first
        ledger.reverse()
        
        return jsonify({
            'partner': partner.to_dict(),
            'ledger': ledger,
            'current_balance': balance
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# --- Excel Export ---
@app.route('/api/products/export', methods=['GET'])
def export_products():
    wb = Workbook()
    ws = wb.active
    ws.title = "Products Export"
    
    # Headers MUST match the Import headers for consistency if possible, or be descriptive
    headers = ['Tên sản phẩm', 'Đơn vị', 'Đơn vị phụ', 'Quy cách', 'Giá vốn', 'Giá bán', 'Tồn kho', 'Hoạt chất']
    ws.append(headers)
    
    products = Product.query.order_by(Product.name).all()
    for p in products:
        ws.append([
            p.name, 
            p.unit, 
            p.secondary_unit or '', 
            p.multiplier, 
            p.cost_price, 
            p.sale_price, 
            p.stock, 
            p.active_ingredient or ''
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='danh_sach_san_pham.xlsx'
    )

@app.route('/api/partners/export', methods=['GET'])
def export_partners():
    wb = Workbook()
    ws = wb.active
    ws.title = "Partners Export"
    
    headers = ['Tên đối tác', 'Loại', 'Số điện thoại', 'Địa chỉ', 'Nợ hiện tại']
    ws.append(headers)
    
    partners = Partner.query.order_by(Partner.name).all()
    for p in partners:
        p_type = 'Khách hàng'
        if p.is_supplier and p.is_customer: p_type = 'Khách & NCC'
        elif p.is_supplier: p_type = 'Nhà cung cấp'
        
        ws.append([
            p.name, 
            p_type, 
            p.phone or '', 
            p.address or '', 
            p.debt_balance
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='danh_sach_doi_tac.xlsx'
    )


# --- Custom Prices (Wholesale) ---
@app.route('/api/custom-prices/<int:partner_id>', methods=['GET'])
def get_custom_prices(partner_id):
    prices = CustomerPrice.query.filter_by(partner_id=partner_id).all()
    # Return as a dict for easy lookup: {product_id: price}
    return jsonify({p.product_id: p.price for p in prices})

@app.route('/api/custom-prices', methods=['POST'])
def save_custom_price():
    data = request.json
    partner_id = data['partner_id']
    product_id = data['product_id']
    price = data['price']
    
    cp = CustomerPrice.query.filter_by(partner_id=partner_id, product_id=product_id).first()
    if not cp:
        cp = CustomerPrice(partner_id=partner_id, product_id=product_id)
        db.session.add(cp)
    
    cp.price = price
    db.session.commit()
    return jsonify(cp.to_dict())

# --- Sales / Purchases (Order) ---
@app.route('/api/orders', methods=['GET'])
def get_orders():
    order_type = request.args.get('type')
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    search_partner = request.args.get('search_partner', '')
    partner_id = request.args.get('partner_id', type=int)
    payment_method = request.args.get('payment_method')
    debt_cycle = request.args.get('debt_cycle', 'false').lower() == 'true'
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product_id = request.args.get('product_id', type=int)
    search_id = request.args.get('search_id', '')
    min_price = request.args.get('minPrice', type=float)
    max_price = request.args.get('maxPrice', type=float)
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)
    
    query = Order.query

    if start_date:
        try:
            dt = datetime.fromisoformat(start_date)
            query = query.filter(Order.date >= dt)
        except: pass
    if end_date:
        try:
            dt = datetime.fromisoformat(end_date)
            if len(end_date) <= 10:
                dt = dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Order.date <= dt)
        except: pass

    if partner_id is not None:
        if partner_id == 0:
            query = query.filter(Order.partner_id == None)
        else:
            query = query.filter(Order.partner_id == partner_id)
        
        if debt_cycle:
            # Simple fallback for debt_cycle if dates not provided
            first_debt_order = Order.query.filter(
                Order.partner_id == (None if partner_id == 0 else partner_id),
                Order.payment_method == 'Debt',
                Order.display_id.notin_(['NODAU', '#NODAU'])
            ).order_by(Order.date.asc()).first()
            if first_debt_order:
                query = query.filter(Order.date >= first_debt_order.date)

    if order_type:
        query = query.filter(Order.type == order_type)
    
    if payment_method:
        query = query.filter(Order.payment_method == payment_method)
    
    if product_id:
        query = query.join(OrderDetail).filter(OrderDetail.product_id == product_id).distinct()
    
    if search_partner:
        s_norm = remove_accents(search_partner)
        query = query.outerjoin(Partner).filter(db.func.remove_accents(db.func.coalesce(Partner.name, 'KHÁCH LẺ')).ilike(f'%{s_norm}%'))
    
    if search_id:
        query = query.filter(Order.display_id.ilike(f'%{search_id}%') | db.cast(Order.id, db.String).ilike(f'%{search_id}%'))
    
    # Price filtering
    if min_price is not None:
        query = query.filter(Order.total_amount >= min_price)
    if max_price is not None:
        query = query.filter(Order.total_amount <= max_price)

    # Date filtering
    if year:
        query = query.filter(db.func.strftime('%Y', Order.date) == str(year))
    if month:
        # Match both 'M' and 'MM' formats from frontend
        m_str = str(month).zfill(2)
        query = query.filter(db.func.strftime('%m', Order.date) == m_str)
    if day:
        d_str = str(day).zfill(2)
        query = query.filter(db.func.strftime('%d', Order.date) == d_str)
    
    quarter = request.args.get('quarter', type=int)
    if quarter:
        # SQLite doesn't have a QUARTER function directly, so we use month ranges
        if quarter == 1:
            query = query.filter(db.func.strftime('%m', Order.date).in_(['01', '02', '03']))
        elif quarter == 2:
            query = query.filter(db.func.strftime('%m', Order.date).in_(['04', '05', '06']))
        elif quarter == 3:
            query = query.filter(db.func.strftime('%m', Order.date).in_(['07', '08', '09']))
        elif quarter == 4:
            query = query.filter(db.func.strftime('%m', Order.date).in_(['10', '11', '12']))
        
    sort_by = request.args.get('sort_by', 'date')
    sort_order = request.args.get('sort_order', 'desc')

    # Mapping frontend sort keys to DB columns/expressions
    sort_map = {
        'id': Order.display_id,
        'date': Order.date,
        'partner_name': db.func.coalesce(Partner.name, 'KHÁCH LẺ'),
        'total_amount': Order.total_amount,
        'payment_method': Order.payment_method
    }

    if sort_by == 'partner_name':
        query = query.outerjoin(Partner)
        
    sort_col = sort_map.get(sort_by, Order.date)
    if sort_order == 'desc':
        query = query.order_by(sort_col.desc())
    else:
        query = query.order_by(sort_col.asc())

    if not partner_id and (not search_id or 'NODAU' not in search_id.upper()):
        query = query.filter(Order.display_id.notin_(['NODAU', '#NODAU']))

    if page and limit:
        # Flask-SQLAlchemy pagination
        pagination = query.paginate(page=page, per_page=limit, error_out=False)
        return jsonify({
            'items': [o.to_dict() for o in pagination.items],
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': pagination.page
        })
    
    orders = query.all()
    return jsonify([o.to_dict() for o in orders])

@app.route('/api/orders', methods=['POST'])
def create_order():
    data = request.json
    # Expected: { partner_id, type: 'Sale'|'Purchase', payment_method: 'Cash'|'Debt', details: [{product_id, quantity, price}] }
    
    try:
        # Custom Order ID Generation (N.DD/MM/YY)
        local_now = datetime.now()
        today_str = local_now.strftime('%d/%m/%y')
        # Filter by date only, ignoring time
        count_today = Order.query.filter(db.func.strftime('%Y-%m-%d', Order.date) == local_now.strftime('%Y-%m-%d')).count()
        display_id = f"{count_today + 1}.{today_str}"
        
        new_order = Order(
            partner_id=data.get('partner_id'),
            type=data['type'],
            payment_method=data['payment_method'],
            display_id=display_id,
            total_amount=0, # will calc
            note=data.get('note'),
            amount_paid=data.get('amount_paid', 0)
        )
        
        total = 0
        for item in data['details']:
            prod = Product.query.get(item['product_id'])
            if not prod:
                raise Exception(f"Product {item['product_id']} not found")
            
            # Inventory Management
            if data['type'] == 'Sale':
                if prod.is_combo:
                    for ci in prod.combo_items:
                        child = Product.query.get(ci.product_id)
                        if child:
                            child.stock = int(child.stock - (item['quantity'] * ci.quantity))
                else:
                    prod.stock = int(prod.stock - item['quantity'])
            elif data['type'] == 'Purchase':
                # Combos are usually not purchased directly, but if they are, we increase child stock
                if prod.is_combo:
                    for ci in prod.combo_items:
                        child = Product.query.get(ci.product_id)
                        if child:
                            child.stock = int(child.stock + (item['quantity'] * ci.quantity))
                else:
                    prod.stock = int(prod.stock + item['quantity'])
                # Update cost_price to the latest purchase price
                prod.cost_price = item['price']
            
            detail = OrderDetail(
                product_id=prod.id,
                product_name_override=item.get('product_name') or item.get('name'), # Support both keys
                quantity=item['quantity'],
                price=item['price']
            )
            new_order.details.append(detail)
            total += item['quantity'] * item['price']
        
        new_order.total_amount = total
        db.session.add(new_order)
        db.session.flush() # ID is now available
        
        # Debt Management & Cash History connection
        if data.get('partner_id'):
            partner = Partner.query.get(data['partner_id'])
            if partner:
                new_order.old_debt = partner.debt_balance
                # NEW LOGIC: Only 'Debt' orders affect balance. 
                # Partial payments at POS do NOT reduce balance or create vouchers.
                if data.get('payment_method') == 'Debt':
                    if data['type'] == 'Sale':
                        partner.debt_balance += total
                        # If total < 0 (Return), amount_paid is what WE give back to CUSTOMER
                        upfront = float(data.get('amount_paid', 0))
                        if upfront > 0:
                            if total >= 0:
                                v_type = 'Receipt'
                                v_note = f"Thanh toán trước cho đơn {display_id}"
                                partner.debt_balance -= upfront
                            else:
                                v_type = 'Payment'
                                v_note = f"Chi trả tiền hàng cho đơn trả {display_id}"
                                partner.debt_balance += upfront
                            
                            v = CashVoucher(
                                partner_id=partner.id,
                                amount=upfront,
                                note=v_note,
                                type=v_type,
                                source='settlement',
                                order_id=new_order.id
                            )
                            db.session.add(v)
                    else:
                        partner.debt_balance -= total
                        # Purchase: upfront payment reduces the negative balance (we pay supplier)
                        # If total < 0 (Return), upfront is what SUPPLIER gives back to US
                        upfront = float(data.get('amount_paid', 0))
                        if upfront > 0:
                            if total >= 0:
                                v_type = 'Payment'
                                v_note = f"Thanh toán trước cho đơn nhập {display_id}"
                                partner.debt_balance += upfront
                            else:
                                v_type = 'Receipt'
                                v_note = f"Thu tiền hàng cho đơn nhập trả {display_id}"
                                partner.debt_balance -= upfront

                            v = CashVoucher(
                                partner_id=partner.id,
                                amount=upfront,
                                note=v_note,
                                type=v_type,
                                source='settlement',
                                order_id=new_order.id
                            )
                            db.session.add(v)
                # Manual vouchers in Fund tab are now the ONLY way to reduce debt.

        # --- Bank Transaction Support ---
        if data.get('payment_method') == 'Transfer' and data.get('bank_account_id'):
            acc_id = int(data['bank_account_id'])
            bank_acc = BankAccount.query.get(acc_id)
            if bank_acc:
                upfront = float(data.get('amount_paid', 0))
                # For Transfer, if amount_paid is 0, we assume the whole total is transferred
                if upfront == 0:
                    upfront = total
                
                t_type = 'Deposit' if data['type'] == 'Sale' else 'Withdrawal'
                # If Sale and total < 0 (Return), it's a Withdrawal from bank
                if data['type'] == 'Sale' and total < 0:
                    t_type = 'Withdrawal'
                # If Purchase and total < 0 (Return), it's a Deposit to bank
                elif data['type'] == 'Purchase' and total < 0:
                    t_type = 'Deposit'

                bt = BankTransaction(
                    account_id=acc_id,
                    amount=abs(upfront),
                    type=t_type,
                    note=f"Thanh toán đơn {display_id}",
                    partner_id=data.get('partner_id'),
                    order_id=new_order.id
                )
                
                if t_type == 'Deposit':
                    bank_acc.balance += abs(upfront)
                else:
                    bank_acc.balance -= abs(upfront)
                
                db.session.add(bt)
                new_order.amount_paid = upfront # Update order state
        
        db.session.commit()
        return jsonify(new_order.to_dict()), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/orders/<int:id>', methods=['GET'])
def get_order(id):
    order = Order.query.get_or_404(id)
    return jsonify(order.to_dict())

@app.route('/api/orders/<int:id>', methods=['DELETE'])
def delete_order(id):
    order = Order.query.get_or_404(id)
    try:
        # 1. Reverse Inventory
        for detail in order.details:
            prod = Product.query.get(detail.product_id)
            if prod:
                if order.type == 'Sale':
                    if prod.is_combo:
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                child.stock = int(child.stock + (detail.quantity * ci.quantity))
                    else:
                        prod.stock = int(prod.stock + detail.quantity)
                elif order.type == 'Purchase':
                    if prod.is_combo:
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                child.stock = int(child.stock - (detail.quantity * ci.quantity))
                    else:
                        prod.stock = int(prod.stock - detail.quantity)
        
        # 2. Reverse Partner Debt
        partner = None
        if order.partner_id:
            partner = Partner.query.get(order.partner_id)
            if partner:
                if order.payment_method == 'Debt':
                    if order.type == 'Sale':
                        partner.debt_balance -= order.total_amount
                    else:
                        partner.debt_balance += order.total_amount
        
        # 3. Cleanup linked settlement vouchers
        linked_vouchers = CashVoucher.query.filter_by(order_id=order.id).all()
        for v in linked_vouchers:
            if partner:
                if v.type == 'Receipt':
                    partner.debt_balance += v.amount
                elif v.type == 'Payment':
                    partner.debt_balance -= v.amount
            db.session.delete(v)
        
        # 4. Cleanup linked bank transactions
        linked_bank_ts = BankTransaction.query.filter_by(order_id=order.id).all()
        for bt in linked_bank_ts:
            bank_acc = BankAccount.query.get(bt.account_id)
            if bank_acc:
                if bt.type == 'Deposit':
                    bank_acc.balance -= bt.amount
                else:
                    bank_acc.balance += bt.amount
            db.session.delete(bt)

        db.session.delete(order)
        db.session.commit()
        return jsonify({'message': 'Order deleted and data reversed successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

def sync_order_amount_paid(order_id):
    """
    Recalculate order.amount_paid based on linked CashVouchers.
    Only applies to Debt orders, as Cash orders do not typically have vouchers.
    """
    order = Order.query.get(order_id)
    if not order: return

    # If it's a Debt order, the amount_paid should be the sum of all settlement vouchers
    if order.payment_method == 'Debt':
        vouchers = CashVoucher.query.filter_by(order_id=order_id).all()
        total_paid = 0
        for v in vouchers:
            # For Sale: Receipt (+) adds to paid amount, Payment (-) refunds it
            if order.type == 'Sale':
                if v.type == 'Receipt': total_paid += v.amount
                elif v.type == 'Payment': total_paid -= v.amount
            # For Purchase: Payment (+) adds to paid amount, Receipt (-) refunds it
            elif order.type == 'Purchase':
                if v.type == 'Payment': total_paid += v.amount
                elif v.type == 'Receipt': total_paid -= v.amount
        
        order.amount_paid = total_paid
        db.session.commit()

@app.route('/api/orders/<int:id>', methods=['PUT'])
def update_order(id):
    order = Order.query.get_or_404(id)
    data = request.json
    try:
        # 1. Reverse Previous Inventory & Debt
        for detail in order.details:
            prod = Product.query.get(detail.product_id)
            if prod:
                if order.type == 'Sale':
                    if prod.is_combo:
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                child.stock = int(child.stock + (detail.quantity * ci.quantity))
                    else:
                        prod.stock = int(prod.stock + detail.quantity)
                elif order.type == 'Purchase':
                    if prod.is_combo:
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                child.stock = int(child.stock - (detail.quantity * ci.quantity))
                    else:
                        prod.stock = int(prod.stock - detail.quantity)
        
        old_debt = 0
        old_partner = None
        if order.partner_id:
            old_partner = Partner.query.get(order.partner_id)
            if old_partner:
                if order.payment_method == 'Debt':
                    if order.type == 'Sale':
                        old_partner.debt_balance -= order.total_amount
                    else:
                        old_partner.debt_balance += order.total_amount
                old_debt = old_partner.debt_balance
        
        # IMPORTANT: If the payment method is changing AWAY from Debt, 
        # we MUST delete associated settlement vouchers because they no longer apply.
        if order.payment_method == 'Debt' and data.get('payment_method') != 'Debt':
            linked_vouchers = CashVoucher.query.filter_by(order_id=order.id).all()
            for v in linked_vouchers:
                if old_partner:
                    if v.type == 'Receipt':
                        old_partner.debt_balance += v.amount
                    else:
                        old_partner.debt_balance -= v.amount
                db.session.delete(v)
        # Reverse Bank Transactions
        old_bank_ts = BankTransaction.query.filter_by(order_id=order.id).all()
        for bt in old_bank_ts:
            bank_acc = BankAccount.query.get(bt.account_id)
            if bank_acc:
                if bt.type == 'Deposit':
                    bank_acc.balance -= bt.amount
                else:
                    bank_acc.balance += bt.amount
            db.session.delete(bt)
            
        # 2. Update Order with New Data
        # Clear existing details
        OrderDetail.query.filter_by(order_id=order.id).delete()
        
        order.partner_id = data.get('partner_id')
        order.payment_method = data['payment_method']
        order.note = data.get('note')
        order.amount_paid = data.get('amount_paid', 0)
        
        if not order.display_id:
            local_now = datetime.now()
            today_str = local_now.strftime('%d/%m/%y')
            today_sql = local_now.strftime('%Y-%m-%d')
            count_today = Order.query.filter(db.func.date(Order.date) == today_sql).count()
            order.display_id = f"{count_today + 1}.{today_str}"
        
        total = 0
        for item in data['details']:
            prod = Product.query.get(item['product_id'])
            if not prod:
                raise Exception(f"Product {item['product_id']} not found")
            
            # Apply New Inventory
            if data['type'] == 'Sale':
                if prod.is_combo:
                    for ci in prod.combo_items:
                        child = Product.query.get(ci.product_id)
                        if child:
                            child.stock = int(child.stock - (item['quantity'] * ci.quantity))
                else:
                    prod.stock = int(prod.stock - item['quantity'])
            elif data['type'] == 'Purchase':
                if prod.is_combo:
                    for ci in prod.combo_items:
                        child = Product.query.get(ci.product_id)
                        if child:
                            child.stock = int(child.stock + (item['quantity'] * ci.quantity))
                else:
                    prod.stock = int(prod.stock + item['quantity'])
                prod.cost_price = item['price']
            
            detail = OrderDetail(
                order_id=order.id,
                product_id=prod.id,
                product_name_override=item.get('product_name') or item.get('name'),
                quantity=item['quantity'],
                price=item['price']
            )
            db.session.add(detail)
            total += item['quantity'] * item['price']
        
        order.total_amount = total
        
        # 3. Apply New Debt
        if order.partner_id:
            partner = Partner.query.get(order.partner_id)
            if partner:
                order.old_debt = partner.debt_balance
                if order.payment_method == 'Debt':
                    if order.type == 'Sale':
                        partner.debt_balance += total
                    else:
                        partner.debt_balance -= total

        # Handle New Bank Transaction
        if data.get('payment_method') == 'Transfer' and data.get('bank_account_id'):
            acc_id = int(data['bank_account_id'])
            bank_acc = BankAccount.query.get(acc_id)
            if bank_acc:
                upfront = float(data.get('amount_paid', 0))
                if upfront == 0:
                    upfront = total
                
                t_type = 'Deposit' if data['type'] == 'Sale' else 'Withdrawal'
                if data['type'] == 'Sale' and total < 0: t_type = 'Withdrawal'
                elif data['type'] == 'Purchase' and total < 0: t_type = 'Deposit'

                bt = BankTransaction(
                    account_id=acc_id,
                    amount=abs(upfront),
                    type=t_type,
                    note=f"Cập nhật đơn {order.display_id}",
                    partner_id=data.get('partner_id'),
                    order_id=order.id
                )
                if t_type == 'Deposit': bank_acc.balance += abs(upfront)
                else: bank_acc.balance -= abs(upfront)
                db.session.add(bt)
                order.amount_paid = upfront

        db.session.commit()
        
        # Enforce consistency: amount_paid must match vouchers
        sync_order_amount_paid(order.id)
        
        order_dict = order.to_dict()
        order_dict['old_debt'] = old_debt
        return jsonify(order_dict)
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# --- Bank Accounts ---
@app.route('/api/bank-accounts', methods=['GET'])
def get_bank_accounts():
    accounts = BankAccount.query.all()
    return jsonify([a.to_dict() for a in accounts])

@app.route('/api/bank-accounts', methods=['POST'])
def create_bank_account():
    data = request.json
    new_acc = BankAccount(
        bank_name=data['bank_name'],
        account_number=data['account_number'],
        account_holder=data.get('account_holder'),
        balance=data.get('balance', 0)
    )
    db.session.add(new_acc)
    db.session.commit()
    return jsonify(new_acc.to_dict()), 201

@app.route('/api/bank-accounts/<int:id>', methods=['PUT'])
def update_bank_account(id):
    acc = BankAccount.query.get_or_404(id)
    data = request.json
    acc.bank_name = data.get('bank_name', acc.bank_name)
    acc.account_number = data.get('account_number', acc.account_number)
    acc.account_holder = data.get('account_holder', acc.account_holder)
    if 'balance' in data:
        acc.balance = float(data['balance'])
    db.session.commit()
    return jsonify(acc.to_dict())

@app.route('/api/bank-accounts/<int:id>', methods=['DELETE'])
def delete_bank_account(id):
    acc = BankAccount.query.get_or_404(id)
    db.session.delete(acc)
    db.session.commit()
    return jsonify({'message': 'Deleted successfully'})

@app.route('/api/bank-transactions', methods=['POST'])
def create_bank_transaction():
    data = request.json
    account_id = data['account_id']
    amount = float(data['amount'])
    t_type = data['type'] # 'Deposit', 'Withdrawal', 'Transfer'
    note = data.get('note', '')
    partner_id = data.get('partner_id')
    
    acc = BankAccount.query.get_or_404(account_id)
    
    transaction = BankTransaction(
        account_id=account_id,
        amount=amount,
        type=t_type,
        note=note,
        partner_id=partner_id
    )
    
    if t_type == 'Deposit':
        acc.balance += amount
    elif t_type == 'Withdrawal':
        acc.balance -= amount
    
    db.session.add(transaction)
    db.session.commit()
    return jsonify(transaction.to_dict()), 201

@app.route('/api/bank-transactions', methods=['GET'])
def get_bank_transactions():
    account_id = request.args.get('account_id', type=int)
    query = BankTransaction.query
    if account_id:
        query = query.filter(BankTransaction.account_id == account_id)
    
    # Optional date filters
    year = request.args.get('year')
    month = request.args.get('month')
    if year:
        query = query.filter(db.func.strftime('%Y', BankTransaction.date) == str(year))
    if month:
        query = query.filter(db.func.strftime('%m', BankTransaction.date) == str(month).zfill(2))
        
    transactions = query.order_by(BankTransaction.date.desc()).all()
    return jsonify([t.to_dict() for t in transactions])

# --- Dashboard ---

@app.route('/api/dashboard-stats', methods=['GET'])
def dashboard_stats():
    from datetime import datetime, timedelta

    # --- 1. Overall Stats ---
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')

    # For revenue and profit, default to today if no filter is provided
    now_dt = datetime.now()
    filter_year = year
    filter_month = month
    filter_day = day

    if not any([year, month, day]):
        filter_year = str(now_dt.year)
        filter_month = str(now_dt.month).zfill(2)
        filter_day = str(now_dt.day).zfill(2)

    # Optimized revenue and profit calculation
    from sqlalchemy import func
    
    # helper for range filter
    def apply_filters(q):
        if filter_year: q = q.filter(db.func.strftime('%Y', Order.date) == str(filter_year))
        if filter_month: q = q.filter(db.func.strftime('%m', Order.date) == str(filter_month).zfill(2))
        if filter_day: q = q.filter(db.func.strftime('%d', Order.date) == str(filter_day).zfill(2))
        return q

    # Revenue is simple sum from Order table (Exclude NODAU)
    revenue = apply_filters(db.session.query(func.sum(Order.total_amount))
                            .filter(Order.type == 'Sale', Order.display_id.notin_(['NODAU', '#NODAU']))).scalar() or 0
    
    # For Profit, we MUST handle combos to match reports.
    # To keep dashboard fast, we'll join Orders with Details and Products.
    # But wait, the combo cost logic in report_products iterates children.
    # In SQL, we can get cost of non-combos + cost of combo components via separate sums or a smart join.
    
    # 1. Simple parts of sales (non-combo)
    cost_simple = apply_filters(db.session.query(func.sum(OrderDetail.quantity * Product.cost_price))
        .join(Order, Order.id == OrderDetail.order_id)
        .join(Product, OrderDetail.product_id == Product.id)
        .filter(Order.type == 'Sale', Product.is_combo == False)).scalar() or 0
    
    # 2. Combo parts
    from sqlalchemy.orm import aliased
    ComponentProduct = aliased(Product)
    ComboProduct = aliased(Product)
    
    cost_combo = apply_filters(db.session.query(func.sum(OrderDetail.quantity * ComboItem.quantity * ComponentProduct.cost_price))
        .join(Order, Order.id == OrderDetail.order_id)
        .join(ComboProduct, OrderDetail.product_id == ComboProduct.id)
        .join(ComboItem, ComboProduct.id == ComboItem.combo_id)
        .join(ComponentProduct, ComboItem.product_id == ComponentProduct.id)
        .filter(Order.type == 'Sale', ComboProduct.is_combo == True)).scalar() or 0

    profit = revenue - (cost_simple + cost_combo)
    
    # --- 2. Debt (Unfiltered - All Time) ---
    all_partners = Partner.query.filter(Partner.debt_balance != 0).all()
    # total_customer_debt: Sum of all positive balances (Receivables) for partners marked as customers
    total_customer_debt = sum(p.debt_balance for p in all_partners if p.is_customer and p.debt_balance > 0)
    # total_supplier_debt: Sum of all absolute negative balances (Payables) for partners marked as suppliers
    total_supplier_debt = sum(abs(p.debt_balance) for p in all_partners if p.is_supplier and p.debt_balance < 0)

    # Lists filtered by type AND appropriate balance sign
    customers_with_debt = sorted([p for p in all_partners if p.is_customer and p.debt_balance > 0], key=lambda x: x.debt_balance, reverse=True)[:10]
    suppliers_with_debt = sorted([p for p in all_partners if p.is_supplier and p.debt_balance < 0], key=lambda x: abs(x.debt_balance), reverse=True)[:10]

    # --- 3. 7-Day Revenue Chart ---
    today_dt = datetime.now()
    seven_days_ago = today_dt - timedelta(days=7)
    
    # Daily Revenue
    daily_revs = db.session.query(
        db.func.date(Order.date).label('day'),
        func.sum(Order.total_amount).label('rev')
    ).filter(Order.type == 'Sale', Order.date >= seven_days_ago, Order.display_id.notin_(['NODAU', '#NODAU']))\
     .group_by(db.func.date(Order.date)).all()
    
    # Daily Cost (Simple)
    daily_costs_simple = db.session.query(
        db.func.date(Order.date).label('day'),
        func.sum(OrderDetail.quantity * Product.cost_price).label('cost')
    ).join(Order, Order.id == OrderDetail.order_id)\
     .join(Product, OrderDetail.product_id == Product.id)\
     .filter(Order.type == 'Sale', Order.date >= seven_days_ago, Product.is_combo == False)\
     .group_by(db.func.date(Order.date)).all()
     
    # Daily Cost (Combo)
    daily_costs_combo = db.session.query(
        db.func.date(Order.date).label('day'),
        func.sum(OrderDetail.quantity * ComboItem.quantity * ComponentProduct.cost_price).label('cost')
    ).join(Order, Order.id == OrderDetail.order_id)\
     .join(ComboProduct, OrderDetail.product_id == ComboProduct.id)\
     .join(ComboItem, ComboProduct.id == ComboItem.combo_id)\
     .join(ComponentProduct, ComboItem.product_id == ComponentProduct.id)\
     .filter(Order.type == 'Sale', Order.date >= seven_days_ago, ComboProduct.is_combo == True)\
     .group_by(db.func.date(Order.date)).all()
    
    rev_map = {str(d.day): d.rev for d in daily_revs}
    cost_map = {str(d.day): d.cost for d in daily_costs_simple}
    for d in daily_costs_combo:
        cost_map[str(d.day)] = cost_map.get(str(d.day), 0) + d.cost
    
    last_7_days = [(today_dt - timedelta(days=i)).date() for i in range(6, -1, -1)] 
    chart_labels = [d.strftime('%d/%m') for d in last_7_days]
    chart_data = []
    chart_profit_data = []

    for d in last_7_days:
        r = rev_map.get(str(d), 0)
        c = cost_map.get(str(d), 0)
        chart_data.append(r)
        chart_profit_data.append(r - c)

    # --- 4. Product Warnings (Expiry & Low Stock) ---
    today = today_dt.date()
    near_expiry_count = 0
    expired_count = 0
    
    # Optimize stock warning: use SQL for counting
    low_stock_count = Product.query.filter(Product.stock < 2 * Product.multiplier).count()
    
    products_with_expiry = Product.query.filter(Product.expiry_date != None).all()
    for p in products_with_expiry:
        # Expiry logic: Near expiry <= 90 days, Expired < 0 days
        try:
            # Handle multiple date formats
            exp_date = None
            date_str = str(p.expiry_date).strip()
            
            # Try parsing commonly used formats
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
                try:
                    exp_date = datetime.strptime(date_str, fmt).date()
                    break
                except ValueError:
                    continue
            
            if exp_date:
                delta = (exp_date - today).days
                
                if delta < 0:
                    expired_count += 1
                elif delta <= 60: # Changed from 90 to 60 as per user request
                    near_expiry_count += 1
        except (ValueError, TypeError):
            pass # Ignore invalid dates

    return jsonify({
        'revenue': revenue,
        'profit': profit,
        'customer_debt': total_customer_debt,
        'supplier_debt': total_supplier_debt,
        'customer_debt_list': [{'id': p.id, 'name': p.name, 'balance': p.debt_balance} for p in customers_with_debt],
        'supplier_debt_list': [{'id': p.id, 'name': p.name, 'balance': p.debt_balance} for p in suppliers_with_debt],
        'chart': {
            'labels': chart_labels,
            'data': chart_data,
            'profit_data': chart_profit_data
        },
        'expiry': {
            'near': near_expiry_count,
            'expired': expired_count
        },
        'low_stock': low_stock_count
    })

def sync_order_amount_paid(order_id):
    if not order_id: return
    order = Order.query.get(order_id)
    if order:
        # Sum all vouchers for this order
        total_paid = db.session.query(db.func.sum(CashVoucher.amount)).filter(CashVoucher.order_id == order_id).scalar() or 0
        order.amount_paid = total_paid
        db.session.commit()

@app.route('/api/vouchers', methods=['POST'])
def create_voucher():
    data = request.json
    partner_id = data.get('partner_id')
    amount = float(data.get('amount', 0))
    note = data.get('note', '')
    v_type = data.get('type', 'Payment')
    
    partner = Partner.query.get(partner_id) if partner_id else None
    
    voucher = CashVoucher(
        partner_id=partner_id,
        amount=amount,
        note=note,
        type=v_type,
        source=data.get('source', 'manual'),
        order_id=data.get('order_id')
    )
    
    if partner and v_type in ['Payment', 'Receipt']:
        # Receipt: Reducing Customer Debt (balance -= amount)
        # Payment: Reducing Supplier Debt (balance += amount towards 0)
        if v_type == 'Receipt':
            partner.debt_balance -= amount
        else:
            partner.debt_balance += amount
        
    db.session.add(voucher)
    db.session.commit()
    
    # Sync amount_paid if linked to an order
    if voucher.order_id:
        sync_order_amount_paid(voucher.order_id)
        
    return jsonify(voucher.to_dict()), 201

@app.route('/api/vouchers', methods=['GET'])
def get_vouchers():
    partner_id = request.args.get('partner_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = CashVoucher.query
    
    source = request.args.get('source')
    if source:
        query = query.filter(CashVoucher.source == source)
    
    if partner_id is not None:
        if partner_id == 0:
            query = query.filter(CashVoucher.partner_id == None)
        else:
            query = query.filter(CashVoucher.partner_id == partner_id)
    
    if start_date:
        try:
            query = query.filter(CashVoucher.date >= datetime.fromisoformat(start_date))
        except: pass
    if end_date:
        try:
            query = query.filter(CashVoucher.date <= datetime.fromisoformat(end_date))
        except: pass

    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    quarter = request.args.get('quarter', type=int)

    if year:
        query = query.filter(db.func.strftime('%Y', CashVoucher.date) == str(year))
    if month:
        m_str = str(month).zfill(2)
        query = query.filter(db.func.strftime('%m', CashVoucher.date) == m_str)
    if day:
        d_str = str(day).zfill(2)
        query = query.filter(db.func.strftime('%d', CashVoucher.date) == d_str)
    
    if quarter:
        if quarter == 1:
            query = query.filter(db.func.strftime('%m', CashVoucher.date).in_(['01', '02', '03']))
        elif quarter == 2:
            query = query.filter(db.func.strftime('%m', CashVoucher.date).in_(['04', '05', '06']))
        elif quarter == 3:
            query = query.filter(db.func.strftime('%m', CashVoucher.date).in_(['07', '08', '09']))
        elif quarter == 4:
            query = query.filter(db.func.strftime('%m', CashVoucher.date).in_(['10', '11', '12']))
        
    vouchers = query.order_by(CashVoucher.date.desc()).all()
    return jsonify([v.to_dict() for v in vouchers])

@app.route('/api/vouchers/<int:id>', methods=['DELETE'])
def delete_voucher(id):
    try:
        voucher = CashVoucher.query.get_or_404(id)
        # Reverse debt change if applicable
        if voucher.partner_id:
            partner = Partner.query.get(voucher.partner_id)
            if partner:
                if voucher.type == 'Receipt':
                    partner.debt_balance += voucher.amount
                elif voucher.type == 'Payment':
                    partner.debt_balance -= voucher.amount
        
        
        # REVERSION LOGIC: If this was a settlement voucher, revert the order to 'Pending'
        if voucher.source == 'settlement' and voucher.order_id:
            order = Order.query.get(voucher.order_id)
            if order:
                order.payment_method = 'Pending'
                order.amount_paid = 0
                # We don't touch inventory/debt balance here because the Voucher deletion 
                # already handled the debt balance reversal above, and moving to 'Pending'
                # doesn't affect stock (it was already subtracted/added during create_order).
        
        db.session.delete(voucher)
        db.session.commit()
        
        # Sync amount_paid if linked to an order
        if voucher.order_id:
            sync_order_amount_paid(voucher.order_id)
            
        return jsonify({'message': 'Voucher deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400



# --- Settings ---
@app.route('/api/print-templates', methods=['GET'])
def get_print_templates():
    module = request.args.get('module')
    if module:
        templates = PrintTemplate.query.filter_by(module=module).all()
    else:
        templates = PrintTemplate.query.all()
    return jsonify([t.to_dict() for t in templates])

@app.route('/api/settings', methods=['GET'])
def get_settings():
    # Return general app settings from AppSetting model
    settings = AppSetting.query.all()
    return jsonify({s.setting_key: s.setting_value for s in settings})

@app.route('/api/settings', methods=['POST'])
def save_setting():
    data = request.json
    for key, value in data.items():
        setting = AppSetting.query.filter_by(setting_key=key).first()
        if setting:
            setting.setting_value = str(value)
        else:
            new_setting = AppSetting(setting_key=key, setting_value=str(value))
            db.session.add(new_setting)
    db.session.commit()
    return jsonify({'message': 'Settings saved successfully'})

@app.route('/api/print-templates', methods=['POST'])
def create_print_template():
    data = request.json
    new_template = PrintTemplate(
        name=data['name'],
        module=data['module'],
        is_default=data.get('is_default', False),
        config=json.dumps(data.get('config', {})),
        content_config=json.dumps(data.get('content_config', {}))
    )
    
    if new_template.is_default:
        # Unset other defaults for this module
        PrintTemplate.query.filter_by(module=new_template.module).update({PrintTemplate.is_default: False})
        
    db.session.add(new_template)
    db.session.commit()
    return jsonify(new_template.to_dict()), 201

@app.route('/api/print-templates/<int:id>', methods=['PUT'])
def update_print_template(id):
    template = PrintTemplate.query.get_or_404(id)
    data = request.json
    template.name = data.get('name', template.name)
    template.module = data.get('module', template.module)
    template.is_default = data.get('is_default', template.is_default)
    if 'config' in data:
        template.config = json.dumps(data['config'])
    if 'content_config' in data:
        template.content_config = json.dumps(data['content_config'])
        
    if template.is_default:
        PrintTemplate.query.filter_by(module=template.module).filter(PrintTemplate.id != id).update({PrintTemplate.is_default: False})
        
    db.session.commit()
    return jsonify(template.to_dict())

@app.route('/api/print-templates/<int:id>', methods=['DELETE'])
def delete_print_template(id):
    template = PrintTemplate.query.get_or_404(id)
    db.session.delete(template)
    db.session.commit()
    return jsonify({'message': 'Template deleted successfully'})

# --- Reports ---
@app.route('/api/reports/products', methods=['GET'])
def report_products():
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    quarter = request.args.get('quarter')
    search = request.args.get('search', '').lower()
    brand = request.args.get('brand')
    sort_by = request.args.get('sort_by', 'revenue')
    sort_order = request.args.get('sort_order', 'desc')
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)
    
    query = OrderDetail.query.join(Order).join(Product).options(joinedload(OrderDetail.product)).filter(Order.type == 'Sale')
    # Date filtering
    if year:
        query = query.filter(db.func.strftime('%Y', Order.date) == str(year))
    if month:
        query = query.filter(db.func.strftime('%m', Order.date) == str(month).zfill(2))
    if quarter:
        quarters = {'1': ('01','03'), '2': ('04','06'), '3': ('07','09'), '4': ('10','12')}
        start, end = quarters.get(quarter, ('01','12'))
        query = query.filter(db.func.strftime('%m', Order.date).between(start, end))
    if day:
        query = query.filter(db.func.strftime('%d', Order.date) == str(day).zfill(2))
    
    if brand:
        query = query.filter(Product.brand == brand)
        
    details = query.all()
    report = {}
    for d in details:
        pid = d.product_id
        if pid not in report:
            p_name = d.product_name_override or (d.product.name if d.product else 'Sản phẩm đã xóa')
            p_unit = d.product.unit if d.product else 'ĐV'
            report[pid] = {
                'id': pid,
                'name': p_name,
                'unit': p_unit,
                'quantity': 0,
                'revenue': 0,
                'cost': 0,
                'profit': 0
            }
        report[pid]['quantity'] += d.quantity
        report[pid]['revenue'] += d.quantity * d.price
        
        # Calculate cost
        current_cost = 0
        if d.product:
            if d.product.is_combo:
                combo_cost = sum(ci.quantity * ci.product.cost_price for ci in d.product.combo_items if ci.product)
                current_cost = d.quantity * combo_cost
            else:
                current_cost = d.quantity * (d.product.cost_price or 0)
        
        report[pid]['cost'] += current_cost
        report[pid]['profit'] = report[pid]['revenue'] - report[pid]['cost']
    
    report_list = list(report.values())
    
    # Search filtering
    if search:
        s_norm = remove_accents(search)
        report_list = [item for item in report_list if s_norm in remove_accents(item['name'])]
    
    # Sorting
    reverse = (sort_order == 'desc')
    def get_sort_key(x):
        val = x.get(sort_by, 0)
        if val is None: return 0
        if isinstance(val, str): return val.lower()
        return val
        
    report_list.sort(key=get_sort_key, reverse=reverse)
    
    total = len(report_list)
    if page and limit:
        start = (page - 1) * limit
        end = start + limit
        items = report_list[start:end]
        return jsonify({
            'items': items,
            'total': total,
            'pages': (total + limit - 1) // limit,
            'current_page': page
        })
    
    return jsonify(report_list)

@app.route('/api/reports/partners', methods=['GET'])
def report_partners():
    p_type = request.args.get('type', 'Customer')
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    quarter = request.args.get('quarter')
    search = request.args.get('search', '').lower()
    sort_by = request.args.get('sort_by', 'total_amount')
    sort_order = request.args.get('sort_order', 'desc')
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)
    
    o_type = 'Sale' if p_type == 'Customer' else 'Purchase'
    query = Order.query.filter(Order.type == o_type, Order.display_id.notin_(['NODAU', '#NODAU']))
    # Date filtering
    if year:
        query = query.filter(db.func.strftime('%Y', Order.date) == str(year))
    if month:
        query = query.filter(db.func.strftime('%m', Order.date) == str(month).zfill(2))
    if quarter:
        quarters = {'1': ('01','03'), '2': ('04','06'), '3': ('07','09'), '4': ('10','12')}
        start, end = quarters.get(quarter, ('01','12'))
        query = query.filter(db.func.strftime('%m', Order.date).between(start, end))
    if day:
        query = query.filter(db.func.strftime('%d', Order.date) == str(day).zfill(2))
        
    orders = query.all()
    report = {}
    for o in orders:
        pid = o.partner_id or 0 # 0 for retail
        pname = o.partner.name if o.partner else ('KHÁCH LẺ' if p_type == 'Customer' else 'NCC VÃNG LAI')
        
        if pid not in report:
            report[pid] = {
                'name': pname,
                'id': pid,
                'count': 0,
                'total_amount': 0,
                'profit': 0 # only relevant for sales
            }
        
        report[pid]['count'] += 1
        report[pid]['total_amount'] += o.total_amount
        
        if o_type == 'Sale':
            order_profit = 0
            for d in o.details:
                if d.product:
                    if d.product.is_combo:
                        combo_cost = sum(ci.quantity * ci.product.cost_price for ci in d.product.combo_items if ci.product)
                        order_profit += d.quantity * (d.price - combo_cost)
                    else:
                        order_profit += d.quantity * (d.price - (d.product.cost_price or 0))
            report[pid]['profit'] += order_profit
            
    report_list = list(report.values())
    
    # Search filtering
    if search:
        s_norm = remove_accents(search)
        report_list = [item for item in report_list if s_norm in remove_accents(item['name'])]
    
    # Sorting
    reverse = (sort_order == 'desc')
    def get_sort_key(x):
        val = x.get(sort_by, 0)
        if val is None: return 0
        if isinstance(val, str): return val.lower()
        return val
        
    report_list.sort(key=get_sort_key, reverse=reverse)
    
    total = len(report_list)
    if page and limit:
        start = (page - 1) * limit
        end = start + limit
        items = report_list[start:end]
        return jsonify({
            'items': items,
            'total': total,
            'pages': (total + limit - 1) // limit,
            'current_page': page
        })
        
    return jsonify(report_list)

@app.route('/api/reports/synthesis', methods=['GET'])
def report_synthesis():
    r_type = request.args.get('type', 'Sale')  # Sale, Purchase
    year = request.args.get('year')
    month = request.args.get('month')
    quarter = request.args.get('quarter')
    day = request.args.get('day')
    # Filter params
    partner_id = request.args.get('partner_id', type=int)
    product_id = request.args.get('product_id', type=int)
    brand = request.args.get('brand')
    
    # Optional Date Range (overrides y/m/d/q if provided)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)

    # Base query
    # We want: Partner Name, Product Name, Brand, Unit, Sum(Qty), Sum(Amount)
    query = db.session.query(
        db.func.coalesce(Partner.name, 'KHÁCH LẺ' if r_type == 'Sale' else 'NCC VÃNG LAI').label('partner_name'),
        Product.name.label('product_name'),
        Product.brand,
        Product.unit,
        db.func.sum(OrderDetail.quantity).label('total_qty'),
        db.func.sum(OrderDetail.quantity * OrderDetail.price).label('total_val'),
        Order.partner_id,
        OrderDetail.product_id
    ).join(Order, OrderDetail.order_id == Order.id)\
     .join(Product, OrderDetail.product_id == Product.id)\
     .outerjoin(Partner, Order.partner_id == Partner.id)

    # Filter by Order Type
    query = query.filter(Order.type == r_type)

    # Date Filtering
    if start_date and end_date:
         query = query.filter(Order.date >= datetime.fromisoformat(start_date))\
                      .filter(Order.date <= datetime.fromisoformat(end_date))
    else:
        if year:
            query = query.filter(db.func.strftime('%Y', Order.date) == str(year))
        if month:
            query = query.filter(db.func.strftime('%m', Order.date) == str(month).zfill(2))
        if day:
            query = query.filter(db.func.strftime('%d', Order.date) == str(day).zfill(2))
        if quarter:
            quarters = {'1': ('01','03'), '2': ('04','06'), '3': ('07','09'), '4': ('10','12')}
            start, end = quarters.get(quarter, ('01','12'))
            query = query.filter(db.func.strftime('%m', Order.date).between(start, end))

    # Detailed Filters
    if partner_id is not None:
        if partner_id == 0:
            query = query.filter(Order.partner_id == None)
        else:
            query = query.filter(Order.partner_id == partner_id)
    if product_id:
        query = query.filter(OrderDetail.product_id == product_id)
    if brand:
        query = query.filter(Product.brand == brand)

    # Group By IDs to be precise
    query = query.group_by(
        Order.partner_id,
        OrderDetail.product_id,
        db.func.coalesce(Partner.name, 'KHÁCH LẺ' if r_type == 'Sale' else 'NCC VÃNG LAI'),
        Product.name,
        Product.brand,
        Product.unit
    )

    # Fetch
    rows = query.all()

    results = []
    for p_name, prod_name, p_brand, unit, qty, val, pid, prod_id in rows:
        results.append({
            'partner_name': p_name,
            'product_name': prod_name,
            'brand': p_brand or '',
            'unit': unit,
            'quantity': qty,
            'revenue': val,
            'partner_id': pid,
            'product_id': prod_id
        })

    # Sorting (handled by frontend mostly, but let's default sort by revenue desc)
    results.sort(key=lambda x: x['revenue'], reverse=True)

    total = len(results)
    if page and limit:
        start = (page - 1) * limit
        end = start + limit
        items = results[start:end]
        return jsonify({
            'items': items,
            'total': total,
            'pages': (total + limit - 1) // limit,
            'current_page': page
        })

    return jsonify(results)

# --- Backup / Restore ---
@app.route('/api/backup', methods=['GET'])
def download_backup():
    try:
        db_path = os.path.join(app.instance_path, 'easypos.db')
        if not os.path.exists(db_path):
             db_path = os.path.join(app.root_path, 'easypos.db')
        
        return send_file(
            db_path,
            as_attachment=True,
            download_name=f"easypos_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/restore', methods=['POST'])
def restore_backup():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    try:
        db_path = get_storage_path(os.path.join("instance", "easypos.db"))
        
        # Close all connections
        db.session.remove()
        db.engine.dispose()
        
        # Save uploaded file over existing db
        file.save(db_path)
        
        # Run migrations on the restored database
        run_migrations()
        
        return jsonify({'message': 'Dữ liệu đã được khôi phục thành công! Hãy khởi động lại ứng dụng để đảm bảo ổn định nhất.'})
    except Exception as e:
        app.logger.error(f"Error restoring backup: {e}")
        return jsonify({'error': str(e)}), 500

# --- Security / Reset ---

# --- Security / Reset ---
@app.route('/api/reset-database', methods=['POST'])
def reset_database():
    data = request.json
    password = data.get('password')
    
    # Simple password check - in a real app this should be more secure
    if password != 'admin.reset':
        return jsonify({'error': 'Sai mật khẩu!'}), 403
        
    try:
        # Delete in order of dependencies (Details first)
        OrderDetail.query.delete()
        Order.query.delete()
        CashVoucher.query.delete()
        CustomerPrice.query.delete()
        ComboItem.query.delete()
        Product.query.delete()
        Partner.query.delete()
        
        db.session.commit()
        return jsonify({'message': 'Đã xóa toàn bộ dữ liệu thành công!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/optimize-db', methods=['POST'])
def optimize_db():
    try:
        with db.engine.begin() as conn:
            conn.execute(db.text('VACUUM'))
        return jsonify({'message': 'Tối ưu hóa dữ liệu thành công!'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db-stats', methods=['GET'])
def db_stats():
    try:
        stats = {
            'orders': Order.query.count(),
            'order_details': OrderDetail.query.count(),
            'products': Product.query.count(),
            'vouchers': CashVoucher.query.count(),
            'partners': Partner.query.count()
        }
        # Get file size
        db_path = get_storage_path(os.path.join("instance", "easypos.db"))
        if os.path.exists(db_path):
            size_mb = os.path.getsize(db_path) / (1024 * 1024)
            stats['db_size_mb'] = round(size_mb, 2)
        else:
            stats['db_size_mb'] = 0
            
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'API Not Found'}), 404
    
    # Safety check for static folder
    if not app.static_folder or not os.path.exists(os.path.join(app.static_folder, 'index.html')):
        return f"Frontend build not found in {app.static_folder}. Please ensure 'frontend/dist' exists.", 500
        
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/api/cash-vouchers', methods=['GET'])
def get_cash_vouchers():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search_partner = request.args.get('search_partner', '')
    voucher_type = request.args.get('type') # Receipt / Payment
    partner_id = request.args.get('partner_id', type=int)
    search_id = request.args.get('search_id', '')

    query = CashVoucher.query
    if search_partner:
        s_norm = remove_accents(search_partner)
        query = query.outerjoin(Partner).filter(db.func.remove_accents(db.func.coalesce(Partner.name, 'Hệ thống')).ilike(f'%{s_norm}%'))

    if start_date:
        try:
            dt = datetime.fromisoformat(start_date)
            query = query.filter(CashVoucher.date >= dt)
        except: pass
    if end_date:
        try:
            dt = datetime.fromisoformat(end_date)
            if len(end_date) <= 10:
                dt = dt.replace(hour=23, minute=59, second=59)
            query = query.filter(CashVoucher.date <= dt)
        except: pass
    
    if partner_id is not None:
        if partner_id == 0:
            query = query.filter(CashVoucher.partner_id == None)
        else:
            query = query.filter(CashVoucher.partner_id == partner_id)

    if search_id:
        # Support searching PT/PC IDs
        query = query.filter(db.cast(CashVoucher.id, db.String).ilike(f'%{search_id}%'))

    if voucher_type:
        query = query.filter(CashVoucher.type == voucher_type)
        
    vouchers = query.order_by(CashVoucher.date.desc()).all()
    return jsonify([v.to_dict() for v in vouchers])

@app.route('/api/ip', methods=['GET'])
def get_ip():
    import socket
    return jsonify({
        'ip': get_local_ip(),
        'port': CURRENT_PORT,
        'hostname': socket.gethostname()
    })

def get_local_ip():
    import socket
    try:
        # Prioritize 192.168 block
        all_ips = socket.gethostbyname_ex(socket.gethostname())[2]
        for ip in all_ips:
            if ip.startswith("192.168."):
                return ip
        
        # Method 2: Connected socket (accurate for internet-connected machines)
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(0.2)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        try:
            # Fallback to any non-loopback
            all_ips = socket.gethostbyname_ex(socket.gethostname())[2]
            for ip in all_ips:
                if not ip.startswith("127."):
                    return ip
            return all_ips[0] if all_ips else "127.0.0.1"
        except:
            return "127.0.0.1"

# --- Serve Frontend ---
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve(path):
    if path != "" and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    else:
        return send_from_directory(app.static_folder, 'index.html')

# --- Integrity Check ---


@app.route('/api/shutdown', methods=['POST'])
def shutdown():
    app.logger.info("Manual shutdown requested. Exiting...")
    # Use a thread to exit so we can return the response first
    threading.Timer(1.0, lambda: os._exit(0)).start()
    return jsonify({'message': 'Server is shutting down...'})

if __name__ == "__main__":
    is_bundle = getattr(sys, 'frozen', False)
    local_ip = get_local_ip()

    try:
        print("\n" + "="*50)
        print("LYANG POS IS RUNNING!")
        print(f"Local access: http://localhost:{CURRENT_PORT}")
        print(f"Network access (Wifi): http://{local_ip}:{CURRENT_PORT}")
        print("="*50 + "\n")
    except:
        pass

    # Start Flask in a background thread
    # Daemon=False so the app stays alive after Splash Screen (Tkinter) closes
    flask_thread = threading.Thread(target=lambda: app.run(host="0.0.0.0", port=CURRENT_PORT, debug=False, use_reloader=False), daemon=False)
    flask_thread.start()
    
    # Show Splash Screen in Main Thread only if not headless
    if not os.environ.get('NO_GUI') and not os.environ.get('HEADLESS'):
        try:
            splash = SplashScreen(CURRENT_PORT)
            splash.run()
        except Exception as e:
            print(f"Could not start Splash Screen: {e}")
            # Keep main thread alive if splash fails or is skipped
            try:
                while True: time.sleep(1)
            except KeyboardInterrupt:
                pass
    else:
        # Headless mode: Keep main thread alive
        try:
            while True: time.sleep(1)
        except KeyboardInterrupt:
            pass

